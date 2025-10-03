#!/usr/bin/env python3
"""
Script to create test-vulnerabilities.xlsx for E2E testing

Related to: Feature 003-i-want-to (Vulnerability Management System)
Task: T030 - Create test Excel file
"""

from openpyxl import Workbook
from openpyxl.styles import Font
import sys

def create_test_vulnerabilities_xlsx(output_path):
    """Create test vulnerabilities Excel file"""

    wb = Workbook()
    ws = wb.active
    ws.title = "Vulnerabilities"

    # Headers
    headers = [
        "Hostname",
        "Local IP",
        "Host groups",
        "Cloud service account ID",
        "Cloud service instance ID",
        "OS version",
        "Active Directory domain",
        "Vulnerability ID",
        "CVSS severity",
        "Vulnerable product versions",
        "Days open"
    ]

    # Make headers bold
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = Font(bold=True)

    # Test data rows
    test_data = [
        # Row 1: MSHome - existing asset (should link to existing asset if it exists)
        [
            "MSHome",
            "192.168.1.10",
            "Production, Database Servers",
            "aws-account-prod-001",
            "i-0abc123def456789",
            "Windows Server 2019",
            "corp.example.com",
            "CVE-2024-0001",
            "9.8 Critical",
            "Microsoft SQL Server 2019 CU1-CU15",
            "45"
        ],

        # Row 2: WebServer01 - first vulnerability (new asset)
        [
            "WebServer01",
            "10.0.1.100",
            "Production, Web Servers",
            "aws-account-prod-002",
            "i-0def456abc789123",
            "Ubuntu 22.04 LTS",
            "",
            "CVE-2024-0002",
            "7.5 High",
            "Apache HTTP Server 2.4.50",
            "30"
        ],

        # Row 3: WebServer01 - second vulnerability (duplicate hostname, different CVE)
        [
            "WebServer01",
            "10.0.1.100",
            "Production, Web Servers",
            "aws-account-prod-002",
            "i-0def456abc789123",
            "Ubuntu 22.04 LTS",
            "",
            "CVE-2024-0003",
            "8.6 High",
            "OpenSSL 1.1.1k",
            "15"
        ],

        # Row 4: NewAsset - minimal data (only hostname and one CVE)
        [
            "NewAsset",
            "",
            "",
            "",
            "",
            "",
            "",
            "CVE-2024-0004",
            "5.3 Medium",
            "",
            ""
        ],

        # Row 5: InvalidRow - missing hostname (should be skipped)
        [
            "",  # Missing hostname
            "192.168.1.200",
            "Test",
            "",
            "",
            "",
            "",
            "CVE-2024-0005",
            "3.1 Low",
            "",
            ""
        ]
    ]

    # Write data rows
    for row_idx, row_data in enumerate(test_data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Adjust column widths for better readability
    column_widths = {
        'A': 20,  # Hostname
        'B': 15,  # Local IP
        'C': 30,  # Host groups
        'D': 25,  # Cloud service account ID
        'E': 25,  # Cloud service instance ID
        'F': 20,  # OS version
        'G': 25,  # Active Directory domain
        'H': 20,  # Vulnerability ID
        'I': 20,  # CVSS severity
        'J': 35,  # Vulnerable product versions
        'K': 12   # Days open
    }

    for col_letter, width in column_widths.items():
        ws.column_dimensions[col_letter].width = width

    # Save workbook
    wb.save(output_path)
    print(f"Created test file: {output_path}")
    print(f"Rows: {len(test_data)} data rows + 1 header row")
    print("Expected import results:")
    print("  - Imported: 4 (rows 1-4)")
    print("  - Skipped: 1 (row 5 - missing hostname)")
    print("  - Assets Created: 2 (WebServer01, NewAsset)")

if __name__ == "__main__":
    output_path = sys.argv[1] if len(sys.argv) > 1 else "test-vulnerabilities.xlsx"

    try:
        create_test_vulnerabilities_xlsx(output_path)
    except ImportError:
        print("ERROR: openpyxl is not installed.")
        print("Install it with: pip install openpyxl")
        sys.exit(1)
