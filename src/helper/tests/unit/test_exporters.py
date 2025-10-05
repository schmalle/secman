"""T041: Unit tests for export formatters."""
import csv
import tempfile
from datetime import datetime, timezone
from pathlib import Path

import pytest
from openpyxl import load_workbook

from src.exporters.csv_exporter import CSVExporter
from src.exporters.txt_exporter import TXTExporter
from src.exporters.xlsx_exporter import XLSXExporter
from src.models.device import Device
from src.models.enums import DeviceType, ExportFormat, Severity
from src.models.export_config import ExportConfiguration
from src.models.vulnerability import Vulnerability
from src.models.vulnerability_record import VulnerabilityRecord


@pytest.fixture
def sample_record():
    """Create a sample vulnerability record for testing."""
    device = Device(
        device_id="device123",
        hostname="WEB-SERVER-01",
        local_ip="10.0.0.1",
        host_groups=["web-servers", "dmz"],
        os_version="Windows Server 2019",
        device_type=DeviceType.SERVER,
        platform_name="Windows",
        ad_domain="CORP.LOCAL",
        cloud_account_id="aws-12345",
        cloud_instance_id="i-abcdef"
    )

    vulnerability = Vulnerability(
        cve_id="CVE-2021-44228",
        severity=Severity.CRITICAL,
        product_versions=["Apache Log4j 2.14.1"],
        days_open=30,
        detected_date=datetime(2023, 1, 1, tzinfo=timezone.utc),
        cvss_score=10.0,
        description="Remote code execution"
    )

    return VulnerabilityRecord(vulnerability=vulnerability, device=device)


@pytest.fixture
def empty_records():
    """Empty record list for testing."""
    return []


class TestXLSXExporter:
    """Test XLSX exporter."""

    def test_export_creates_file(self, sample_record, tmp_path):
        """XLSX file should be created."""
        config = ExportConfiguration(format=ExportFormat.XLSX)
        exporter = XLSXExporter(config)
        output_file = tmp_path / "test.xlsx"

        exporter.export([sample_record], str(output_file))

        assert output_file.exists()

    def test_export_column_order(self, sample_record, tmp_path):
        """XLSX should have correct column order."""
        config = ExportConfiguration(format=ExportFormat.XLSX)
        exporter = XLSXExporter(config)
        output_file = tmp_path / "test.xlsx"

        exporter.export([sample_record], str(output_file))

        # Load and verify
        wb = load_workbook(str(output_file))
        ws = wb.active
        headers = [cell.value for cell in ws[1]]

        expected_headers = [
            "Hostname", "Local IP", "Host groups",
            "Cloud service account ID", "Cloud service instance ID",
            "OS version", "Active Directory domain",
            "Vulnerability ID", "CVSS severity",
            "Vulnerable product versions", "Days open"
        ]
        assert headers == expected_headers

    def test_export_data_values(self, sample_record, tmp_path):
        """XLSX should contain correct data values."""
        config = ExportConfiguration(format=ExportFormat.XLSX)
        exporter = XLSXExporter(config)
        output_file = tmp_path / "test.xlsx"

        exporter.export([sample_record], str(output_file))

        # Load and verify
        wb = load_workbook(str(output_file))
        ws = wb.active
        row_values = [cell.value for cell in ws[2]]

        assert row_values[0] == "WEB-SERVER-01"  # Hostname
        assert row_values[1] == "10.0.0.1"  # Local IP
        assert row_values[7] == "CVE-2021-44228"  # Vulnerability ID
        assert row_values[8] == "CRITICAL"  # Severity

    def test_export_empty_results(self, tmp_path):
        """XLSX should handle empty results (headers only)."""
        config = ExportConfiguration(format=ExportFormat.XLSX)
        exporter = XLSXExporter(config)
        output_file = tmp_path / "test.xlsx"

        exporter.export([], str(output_file))

        # Load and verify
        wb = load_workbook(str(output_file))
        ws = wb.active
        assert ws.max_row == 1  # Only header row


class TestCSVExporter:
    """Test CSV exporter."""

    def test_export_creates_file(self, sample_record, tmp_path):
        """CSV file should be created."""
        config = ExportConfiguration(format=ExportFormat.CSV)
        exporter = CSVExporter(config)
        output_file = tmp_path / "test.csv"

        exporter.export([sample_record], str(output_file))

        assert output_file.exists()

    def test_export_column_order(self, sample_record, tmp_path):
        """CSV should have correct column order."""
        config = ExportConfiguration(format=ExportFormat.CSV)
        exporter = CSVExporter(config)
        output_file = tmp_path / "test.csv"

        exporter.export([sample_record], str(output_file))

        # Read and verify
        with open(output_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            headers = next(reader)

        expected_headers = [
            "Hostname", "Local IP", "Host groups",
            "Cloud service account ID", "Cloud service instance ID",
            "OS version", "Active Directory domain",
            "Vulnerability ID", "CVSS severity",
            "Vulnerable product versions", "Days open"
        ]
        assert headers == expected_headers

    def test_export_comma_delimited(self, sample_record, tmp_path):
        """CSV should be comma-delimited."""
        config = ExportConfiguration(format=ExportFormat.CSV)
        exporter = CSVExporter(config)
        output_file = tmp_path / "test.csv"

        exporter.export([sample_record], str(output_file))

        # Read raw content
        with open(output_file, 'r', encoding='utf-8') as f:
            first_line = f.readline()

        assert ',' in first_line

    def test_export_proper_quoting(self, tmp_path):
        """CSV should properly quote fields containing commas."""
        # Create record with comma in hostname
        device = Device(
            device_id="device123",
            hostname="SERVER, BACKUP",
            local_ip="10.0.0.1",
            host_groups=["servers"],
            os_version="Linux",
            device_type=DeviceType.SERVER,
            platform_name="Linux"
        )
        vulnerability = Vulnerability(
            cve_id="CVE-2021-12345",
            severity=Severity.HIGH,
            product_versions=["Test"],
            days_open=10,
            detected_date=datetime.now(timezone.utc)
        )
        record = VulnerabilityRecord(vulnerability=vulnerability, device=device)

        config = ExportConfiguration(format=ExportFormat.CSV)
        exporter = CSVExporter(config)
        output_file = tmp_path / "test.csv"

        exporter.export([record], str(output_file))

        # Read with csv reader
        with open(output_file, 'r', encoding='utf-8') as f:
            reader = csv.reader(f)
            next(reader)  # Skip header
            row = next(reader)

        # Hostname should be parsed correctly despite comma
        assert row[0] == "SERVER, BACKUP"

    def test_export_empty_results(self, tmp_path):
        """CSV should handle empty results (headers only)."""
        config = ExportConfiguration(format=ExportFormat.CSV)
        exporter = CSVExporter(config)
        output_file = tmp_path / "test.csv"

        exporter.export([], str(output_file))

        # Read and verify
        with open(output_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        assert len(lines) == 1  # Only header


class TestTXTExporter:
    """Test TXT exporter."""

    def test_export_creates_file(self, sample_record, tmp_path):
        """TXT file should be created."""
        config = ExportConfiguration(format=ExportFormat.TXT)
        exporter = TXTExporter(config)
        output_file = tmp_path / "test.txt"

        exporter.export([sample_record], str(output_file))

        assert output_file.exists()

    def test_export_tab_delimited(self, sample_record, tmp_path):
        """TXT should be tab-delimited."""
        config = ExportConfiguration(format=ExportFormat.TXT)
        exporter = TXTExporter(config)
        output_file = tmp_path / "test.txt"

        exporter.export([sample_record], str(output_file))

        # Read raw content
        with open(output_file, 'r', encoding='utf-8') as f:
            first_line = f.readline()

        assert '\t' in first_line

    def test_export_column_structure(self, sample_record, tmp_path):
        """TXT should match CSV/XLSX column structure."""
        config = ExportConfiguration(format=ExportFormat.TXT)
        exporter = TXTExporter(config)
        output_file = tmp_path / "test.txt"

        exporter.export([sample_record], str(output_file))

        # Read and verify
        with open(output_file, 'r', encoding='utf-8') as f:
            headers = f.readline().strip().split('\t')

        expected_headers = [
            "Hostname", "Local IP", "Host groups",
            "Cloud service account ID", "Cloud service instance ID",
            "OS version", "Active Directory domain",
            "Vulnerability ID", "CVSS severity",
            "Vulnerable product versions", "Days open"
        ]
        assert headers == expected_headers

    def test_export_empty_results(self, tmp_path):
        """TXT should handle empty results (headers only)."""
        config = ExportConfiguration(format=ExportFormat.TXT)
        exporter = TXTExporter(config)
        output_file = tmp_path / "test.txt"

        exporter.export([], str(output_file))

        # Read and verify
        with open(output_file, 'r', encoding='utf-8') as f:
            lines = f.readlines()

        assert len(lines) == 1  # Only header
