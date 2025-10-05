"""T030: XLSX exporter using openpyxl."""
from pathlib import Path
from typing import List

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from src.models.export_config import ExportConfiguration
from src.models.vulnerability_record import VulnerabilityRecord


class XLSXExporter:
    """Export vulnerability records to Excel format."""

    def __init__(self, config: ExportConfiguration) -> None:
        """Initialize XLSX exporter."""
        self.config = config

    def export(self, records: List[VulnerabilityRecord], output_path: str) -> None:
        """Export records to XLSX file."""
        wb = Workbook()
        ws = wb.active
        assert ws is not None, "Active worksheet should not be None"
        ws.title = "Vulnerabilities"

        # Write header row
        headers = self.config.column_order
        for col_idx, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_idx, value=header)

        # Write data rows
        for row_idx, record in enumerate(records, start=2):
            row_data = record.to_row()
            for col_idx, value in enumerate(row_data, start=1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Save workbook
        wb.save(output_path)
