#!/usr/bin/env python3
"""
Generate test Excel files for UserMapping import feature
Feature: 013-user-mapping-upload
"""

import openpyxl
from pathlib import Path

def create_excel_file(filename, headers, data):
    """Create an Excel file with given headers and data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    for col_idx, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col_idx, value=header)
    
    # Add data rows
    for row_idx, row_data in enumerate(data, start=2):
        for col_idx, value in enumerate(row_data, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)
    
    wb.save(filename)
    print(f"✓ Created: {filename}")

# Base directory
base_dir = Path("/Users/flake/sources/misc/secman/src/backendng/src/test/resources/user-mapping-test-files")

# 1. Valid file - 5 valid rows
create_excel_file(
    base_dir / "user-mappings-valid.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["john@example.com", "123456789012", "example.com"],
        ["jane@example.com", "987654321098", "example.com"],
        ["admin@corp.com", "111111111111", "corp.com"],
        ["consultant@agency.com", "555555555555", "clientA.com"],
        ["user@multi.com", "222222222222", "multi.com"],
    ]
)

# 2. Invalid email - 1 invalid, 1 valid
create_excel_file(
    base_dir / "user-mappings-invalid-email.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["notanemail", "123456789012", "example.com"],  # Invalid - no @
        ["valid@example.com", "987654321098", "example.com"],  # Valid
    ]
)

# 3. Invalid AWS account (too short)
create_excel_file(
    base_dir / "user-mappings-invalid-aws.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["john@example.com", "12345", "example.com"],  # Invalid - too short
        ["jane@example.com", "987654321098", "example.com"],  # Valid
    ]
)

# 4. Invalid AWS account (non-numeric)
create_excel_file(
    base_dir / "user-mappings-invalid-aws-nonnumeric.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["john@example.com", "ABC123456789", "example.com"],  # Invalid - non-numeric
        ["jane@example.com", "987654321098", "example.com"],  # Valid
    ]
)

# 5. Invalid domain (contains space)
create_excel_file(
    base_dir / "user-mappings-invalid-domain.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["john@example.com", "123456789012", "example .com"],  # Invalid - space
        ["jane@example.com", "987654321098", "example.com"],  # Valid
    ]
)

# 6. Duplicates - 3 rows with 1 duplicate
create_excel_file(
    base_dir / "user-mappings-duplicates.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["john@example.com", "123456789012", "example.com"],
        ["jane@example.com", "987654321098", "example.com"],
        ["john@example.com", "123456789012", "example.com"],  # Duplicate of row 1
    ]
)

# 7. Missing column (no Domain column)
create_excel_file(
    base_dir / "user-mappings-missing-column.xlsx",
    ["Email Address", "AWS Account ID"],  # Missing "Domain"
    [
        ["john@example.com", "123456789012"],
        ["jane@example.com", "987654321098"],
    ]
)

# 8. Empty file (only headers, no data)
create_excel_file(
    base_dir / "user-mappings-empty.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    []  # No data rows
)

# 9. Mixed valid/invalid for comprehensive testing
create_excel_file(
    base_dir / "user-mappings-mixed.xlsx",
    ["Email Address", "AWS Account ID", "Domain"],
    [
        ["valid1@example.com", "123456789012", "example.com"],  # Valid
        ["notanemail", "987654321098", "example.com"],  # Invalid email
        ["valid2@example.com", "12345", "example.com"],  # Invalid AWS account
        ["valid3@example.com", "111111111111", "bad domain"],  # Invalid domain (space)
        ["valid4@example.com", "222222222222", "example.com"],  # Valid
    ]
)

print("\n✅ All test Excel files created successfully!")
print(f"📁 Location: {base_dir}")
