#!/usr/bin/env python3
"""
Generate E2E test files for user mapping upload
Feature: 013-user-mapping-upload
Reference: specs/013-user-mapping-upload/tasks.md T018
"""

import openpyxl
from pathlib import Path

# Output directory
output_dir = Path("/Users/flake/sources/misc/secman/testdata/user-mappings")
output_dir.mkdir(parents=True, exist_ok=True)

def create_workbook_with_data(data, headers=None):
    """Create workbook with given data"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Add headers
    if headers is None:
        headers = ["Email Address", "AWS Account ID", "Domain"]
    
    for idx, header in enumerate(headers, start=1):
        ws.cell(1, idx, header)
        ws.cell(1, idx).font = openpyxl.styles.Font(bold=True)
    
    # Add data
    for row_idx, row in enumerate(data, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row_idx, col_idx, value)
    
    return wb

# 1. valid-mappings.xlsx - Valid data for successful import
print("Creating valid-mappings.xlsx...")
wb = create_workbook_with_data([
    ["user1@example.com", "123456789012", "example.com"],
    ["user2@example.com", "987654321098", "example.com"],
    ["admin@company.org", "555555555555", "company.org"],
])
wb.save(output_dir / "valid-mappings.xlsx")

# 2. invalid-emails.xlsx - Invalid email formats
print("Creating invalid-emails.xlsx...")
wb = create_workbook_with_data([
    ["notanemail", "123456789012", "example.com"],  # Missing @
    ["user@", "987654321098", "example.com"],  # Missing domain
    ["@example.com", "555555555555", "example.com"],  # Missing local part
    ["user with space@example.com", "111111111111", "example.com"],  # Space in email
])
wb.save(output_dir / "invalid-emails.xlsx")

# 3. invalid-aws-accounts.xlsx - Invalid AWS account IDs
print("Creating invalid-aws-accounts.xlsx...")
wb = create_workbook_with_data([
    ["user1@example.com", "12345", "example.com"],  # Too short
    ["user2@example.com", "1234567890123", "example.com"],  # Too long
    ["user3@example.com", "ABC123456789", "example.com"],  # Not numeric
    ["user4@example.com", "123456789012X", "example.com"],  # Has letter
    ["user5@example.com", "12345 789012", "example.com"],  # Has space
])
wb.save(output_dir / "invalid-aws-accounts.xlsx")

# 4. invalid-domains.xlsx - Invalid domain formats
print("Creating invalid-domains.xlsx...")
wb = create_workbook_with_data([
    ["user1@example.com", "123456789012", "bad domain"],  # Has space
    ["user2@example.com", "987654321098", "domain@bad"],  # Has @
    ["user3@example.com", "555555555555", "domain/bad"],  # Has /
    ["user4@example.com", "111111111111", "UPPERCASE.COM"],  # Uppercase (will be normalized)
])
wb.save(output_dir / "invalid-domains.xlsx")

# 5. missing-columns.xlsx - Missing required columns
print("Creating missing-columns.xlsx...")
wb = create_workbook_with_data(
    [["user1@example.com", "123456789012"]],  # Missing Domain column
    headers=["Email Address", "AWS Account ID"]
)
wb.save(output_dir / "missing-columns.xlsx")

# 6. empty-file.xlsx - Empty file with headers only
print("Creating empty-file.xlsx...")
wb = create_workbook_with_data([])  # No data rows
wb.save(output_dir / "empty-file.xlsx")

# 7. mixed-valid-invalid.xlsx - Mix of valid and invalid rows
print("Creating mixed-valid-invalid.xlsx...")
wb = create_workbook_with_data([
    ["valid1@example.com", "123456789012", "example.com"],  # Valid
    ["notanemail", "987654321098", "example.com"],  # Invalid email
    ["valid2@example.com", "555555555555", "example.com"],  # Valid
    ["valid3@example.com", "12345", "example.com"],  # Invalid AWS ID
    ["valid4@example.com", "111111111111", "example.com"],  # Valid
])
wb.save(output_dir / "mixed-valid-invalid.xlsx")

# 8. duplicates.xlsx - Duplicate mappings
print("Creating duplicates.xlsx...")
wb = create_workbook_with_data([
    ["user1@example.com", "123456789012", "example.com"],  # First occurrence
    ["user1@example.com", "123456789012", "example.com"],  # Duplicate
    ["user2@example.com", "987654321098", "example.com"],  # Different
    ["user1@example.com", "123456789012", "example.com"],  # Another duplicate
])
wb.save(output_dir / "duplicates.xlsx")

# 9. large-file.xlsx - Large file with 150 valid rows
print("Creating large-file.xlsx...")
data = []
for i in range(1, 151):
    email = f"user{i}@example.com"
    aws_id = f"{i:012d}"  # Format as 12-digit number
    domain = f"domain{i % 10}.com"  # Cycle through 10 domains
    data.append([email, aws_id, domain])
wb = create_workbook_with_data(data)
wb.save(output_dir / "large-file.xlsx")

# 10. special-characters.xlsx - Edge cases with special characters
print("Creating special-characters.xlsx...")
wb = create_workbook_with_data([
    ["user+tag@example.com", "123456789012", "example.com"],  # + in email
    ["user.name@example.co.uk", "987654321098", "example.co.uk"],  # Multiple dots
    ["user_name@example.com", "555555555555", "sub-domain.example.com"],  # Underscore and hyphen
    ["user@subdomain.example.org", "111111111111", "example-org.com"],  # Subdomain
])
wb.save(output_dir / "special-characters.xlsx")

# 11. wrong-format.txt - Wrong file format (not Excel)
print("Creating wrong-format.txt...")
with open(output_dir / "wrong-format.txt", "w") as f:
    f.write("Email Address,AWS Account ID,Domain\n")
    f.write("user1@example.com,123456789012,example.com\n")

# 12. empty-cells.xlsx - Files with empty cells
print("Creating empty-cells.xlsx...")
wb = create_workbook_with_data([
    ["user1@example.com", "123456789012", "example.com"],  # Valid
    ["", "987654321098", "example.com"],  # Empty email
    ["user2@example.com", "", "example.com"],  # Empty AWS ID
    ["user3@example.com", "555555555555", ""],  # Empty domain
    ["", "", ""],  # All empty
])
wb.save(output_dir / "empty-cells.xlsx")

# 13. Create a summary file
summary_file = output_dir / "README.md"
with open(summary_file, "w") as f:
    f.write("# User Mapping Upload - E2E Test Data Files\n\n")
    f.write("Generated for Feature 013-user-mapping-upload\n\n")
    f.write("## Test Files\n\n")
    f.write("| File | Purpose | Expected Behavior |\n")
    f.write("|------|---------|-------------------|\n")
    f.write("| valid-mappings.xlsx | 3 valid mappings | Import all 3 successfully |\n")
    f.write("| invalid-emails.xlsx | 4 invalid email formats | Skip all with error messages |\n")
    f.write("| invalid-aws-accounts.xlsx | 5 invalid AWS account IDs | Skip all with error messages |\n")
    f.write("| invalid-domains.xlsx | 4 invalid domain formats | Skip 3, normalize 1 uppercase |\n")
    f.write("| missing-columns.xlsx | Missing Domain column | Fail with header error |\n")
    f.write("| empty-file.xlsx | Headers only, no data | Import 0, skip 0 |\n")
    f.write("| mixed-valid-invalid.xlsx | 3 valid, 2 invalid | Import 3, skip 2 with errors |\n")
    f.write("| duplicates.xlsx | 2 unique, 2 duplicates | Import 2, skip 2 duplicates |\n")
    f.write("| large-file.xlsx | 150 valid rows | Import all 150 successfully |\n")
    f.write("| special-characters.xlsx | Valid edge cases | Import all 4 successfully |\n")
    f.write("| wrong-format.txt | Text file, not Excel | Fail with format error |\n")
    f.write("| empty-cells.xlsx | Some empty cells | Skip 4 rows with missing data |\n")
    f.write("\n## Usage\n\n")
    f.write("```bash\n")
    f.write("# Generate files\n")
    f.write("python3 scripts/generate_e2e_test_files.py\n\n")
    f.write("# Run E2E tests\n")
    f.write("npm test -- user-mapping-upload.spec.ts\n")
    f.write("```\n")

print(f"\n✓ Created 12 test files in {output_dir}")
print(f"✓ Created README.md")
print("\nTest files summary:")
print("  - valid-mappings.xlsx (3 rows)")
print("  - invalid-emails.xlsx (4 rows)")
print("  - invalid-aws-accounts.xlsx (5 rows)")
print("  - invalid-domains.xlsx (4 rows)")
print("  - missing-columns.xlsx (1 row, missing column)")
print("  - empty-file.xlsx (0 rows)")
print("  - mixed-valid-invalid.xlsx (5 rows)")
print("  - duplicates.xlsx (4 rows)")
print("  - large-file.xlsx (150 rows)")
print("  - special-characters.xlsx (4 rows)")
print("  - wrong-format.txt (text file)")
print("  - empty-cells.xlsx (5 rows with empty cells)")
