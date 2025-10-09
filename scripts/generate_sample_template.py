#!/usr/bin/env python3
"""
Generate sample Excel template for user mapping upload
Feature: 013-user-mapping-upload
"""

import openpyxl
from pathlib import Path

# Output file path
output_file = Path("/Users/flake/sources/misc/secman/src/frontend/public/sample-files/user-mapping-template.xlsx")

# Create workbook
wb = openpyxl.Workbook()

# Sheet 1: Mappings
ws1 = wb.active
ws1.title = "Mappings"

# Headers
ws1['A1'] = "Email Address"
ws1['B1'] = "AWS Account ID"
ws1['C1'] = "Domain"

# Make headers bold
for cell in ws1['1']:
    cell.font = openpyxl.styles.Font(bold=True)
    cell.fill = openpyxl.styles.PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid")

# Sample data
data = [
    ["user1@example.com", "123456789012", "example.com"],
    ["user2@example.com", "987654321098", "example.com"],
    ["consultant@agency.com", "555555555555", "clientA.com"],
]

for idx, row in enumerate(data, start=2):
    ws1[f'A{idx}'] = row[0]
    ws1[f'B{idx}'] = row[1]
    ws1[f'C{idx}'] = row[2]

# Set column widths
ws1.column_dimensions['A'].width = 30
ws1.column_dimensions['B'].width = 18
ws1.column_dimensions['C'].width = 25

# Sheet 2: Instructions
ws2 = wb.create_sheet("Instructions")

instructions = [
    ["USER MAPPING TEMPLATE INSTRUCTIONS", ""],
    ["", ""],
    ["Purpose", "This template is used to upload user-to-AWS-account-to-domain mappings for role-based access control."],
    ["", ""],
    ["Required Columns", ""],
    ["1. Email Address", "User's email address (must contain @)"],
    ["2. AWS Account ID", "12-digit numeric AWS account identifier"],
    ["3. Domain", "Organizational domain name (alphanumeric + dots + hyphens)"],
    ["", ""],
    ["Format Rules", ""],
    ["Email", "Must be valid email format (e.g., user@example.com)"],
    ["AWS Account ID", "Exactly 12 numeric digits (e.g., 123456789012)"],
    ["Domain", "Lowercase alphanumeric characters, dots, and hyphens (e.g., example.com)"],
    ["", ""],
    ["Notes", ""],
    ["• One mapping per row", ""],
    ["• Same email can have multiple AWS accounts and domains", ""],
    ["• Duplicate mappings will be automatically skipped", ""],
    ["• Invalid rows will be skipped with error details", ""],
    ["• Maximum file size: 10 MB", ""],
    ["• Only .xlsx (Excel 2007+) format is supported", ""],
    ["", ""],
    ["Examples", ""],
    ["Valid", "john@example.com, 123456789012, example.com"],
    ["Valid", "john@example.com, 987654321098, example.com (same user, different AWS account)"],
    ["Valid", "john@example.com, 123456789012, other.com (same user, different domain)"],
    ["Invalid", "notanemail, 123456789012, example.com (missing @ in email)"],
    ["Invalid", "john@example.com, 12345, example.com (AWS account too short)"],
    ["Invalid", "john@example.com, ABC123456789, example.com (AWS account not numeric)"],
    ["Invalid", "john@example.com, 123456789012, bad domain (domain contains space)"],
]

for row_idx, (col1, col2) in enumerate(instructions, start=1):
    ws2[f'A{row_idx}'] = col1
    ws2[f'B{row_idx}'] = col2

# Style title
ws2['A1'].font = openpyxl.styles.Font(bold=True, size=14)

# Style section headers
for row_idx in [5, 10, 15, 23]:
    ws2[f'A{row_idx}'].font = openpyxl.styles.Font(bold=True)

# Set column widths
ws2.column_dimensions['A'].width = 25
ws2.column_dimensions['B'].width = 70

# Save workbook
wb.save(output_file)
print(f"✓ Created: {output_file}")
print(f"  Size: {output_file.stat().st_size / 1024:.1f} KB")
